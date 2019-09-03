from django.shortcuts import render
from django.http import HttpResponse, HttpResponseRedirect
from .models import *
from . import forms
from django.forms import modelformset_factory
import urllib.parse as ap
import urllib.request
from django.contrib.auth.hashers import make_password, check_password
import random
from django.db.models import Q
from itertools import chain
from openpyxl import Workbook
from django.http import HttpResponse, HttpResponseRedirect
from openpyxl.writer.excel import save_virtual_workbook
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.forms import formset_factory
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

# from selenium import webdriver
# import pdfkit
from django.core.mail import send_mail, EmailMessage



# def email_doc(self, name):
# 	qs = User.objects.get(username=name)
# 	url = 'https://pbas.bmsit.ac.in/ao_teacher1_display/' + name


# 	# DRIVER = 'chromedriver'
# 	# driver = webdriver.Chrome(DRIVER)
# 	# driver.get(url)
# 	# scn = name + '.png'
# 	# # screenshot = driver.get_screenshot_as_png()
# 	# screenshot = driver.save_screenshot(scn)
# 	# driver.quit()


# 	# pdf = pdfkit.from_url(url, False)

# 	email = EmailMessage(
# 				'PBAS',
# 				'Hi, ' + qs.first_name + '\n\n' +'Your PBAS report is attached below: \n\nThanks,\nAO-BMSIT' ,
# 				'PBAS Support <ao@bmsit.in>',
# 				[qs.email],
# 				)
# 	email.attach_file(screenshot)
# 	email.send()

def ao_principal(request):
	if request.method == 'POST':
		y=forms.yearform(request.POST)
		year=y.save(commit=False)
		print(year.year)
		if request.user.is_ao():
			return HttpResponseRedirect(reverse('facform1:ao_first',args=(year.year,)))
		if request.user.is_principal():
			return HttpResponseRedirect(reverse('facform1:principal_first',args=(year.year,)))
	y=forms.yearform()
	return render(request,'ao_principal.html',{'y':y})



def first_page(request):
	user=User.objects.get(username=request.user)
	print("                                                             Entered first_page  try section")
	if request.method == 'POST' and user.profile_pic:
		form=forms.newform(data=request.POST)
		if form.is_valid():
			f=form.save(commit=False)
			name = request.user
			user.designation=f.designation
			user.save()
			if new.objects.filter(info=request.user).filter(year=f.year).exists():
				if user.is_assistant_professor():
					print("User is assistant professor")
					return HttpResponseRedirect(reverse('facform1:assistant_form1',args=(f.year,)))

				elif user.is_associate_professor():
					print("user is associate professor")
					return HttpResponseRedirect(reverse('facform1:associate_form1',args=(f.year,)))

				elif user.is_professor():
					print("user is professor")
					return HttpResponseRedirect(reverse('facform1:associate_form1',args=(f.year,)))
				elif user.is_hod():
					print("User is HOD")
					return HttpResponseRedirect(reverse('facform1:hod_first',args=(f.year,)))
				return HttpResponse("YOU ARE NOT AO OR PRINCIPAL,PLEASE ENTER YOUR CURRENT DESIGNATION")

			else:
				print("                                                             Entered first_page except section")
				if request.method == 'POST':
					form=forms.newform(data=request.POST)
					if form.is_valid():
						f=form.save(commit=False)
						f.info=request.user
						name = request.user
						u = User.objects.get(username=name)
						u.designation=f.designation
						u.save()
						f.save()
						if request.user.is_assistant_professor():
							return HttpResponseRedirect(reverse('facform1:assistant_form1',args=(f.year,)))

						elif request.user.is_associate_professor():
							return HttpResponseRedirect(reverse('facform1:associate_form1',args=(f.year,)))

						elif request.user.is_professor():
							return HttpResponseRedirect(reverse('facform1:associate_form1',args=(f.year,)))
						elif request.user.is_hod():
							return HttpResponseRedirect(reverse('facform1:hod_first',args=(f.year,)))
						return HttpResponse("YOU ARE NOT AO OR PRINCIPAL,PLEASE ENTER YOUR CURRENT DESIGNATION")


					else:
						print(form.errors)

				form=forms.newform()
				return render(request,'first_page.html',{'form':form,'user':user})

		else:
			print(form.errors)
	elif request.method == "POST" and not user.profile_pic:
		print("elif")
		f=request.FILES['abc']
		print(f)
		user.profile_pic = f
		user.save()
		"""
		f=forms.profileform(request.POST,request.FILES)
		if f.is_valid():
			user.profile_pic=f.save(commit=False)
			print(f)
			user.save()
			return HttpResponseRedirect(reverse('facform1:first_page'))
		else:
			HttpResponse("INVALID FILES UPLOADED")"""
		return HttpResponseRedirect(reverse('facform1:first_page'))
	elif not user.profile_pic:
		return render(request,'first_page.html',{'user':user })
	else:
		form=forms.newform()
		return render(request,'first_page.html',{'form':form,'user':user})



"""
def teach_fun(request):

	data1 = User.objects.all()
	for i in data1:
		i.teach_status = False
		i.hod_status = False
		i.principal_status = False
		i.save()

	return HttpResponse("Done")
"""


def report(request,dept):

	data1 = rd.objects.filter(info__department__name=dept)
	data2 = journal.objects.filter(info__department__name=dept)
	data3 = conference.objects.filter(info__department__name=dept)

	book = Workbook()
	book.create_sheet('Rd')
	book.create_sheet('Journal')
	book.create_sheet('Conference')



	rd1 = book['Rd']
	journal1 = book['Journal']
	conference1 = book['Conference']


	#-------------------rd-------------------

	rd1['A1'] = "Name"
	rd1['B1'] = "Department"
	rd1['C1'] = "Designation"
	rd1['D1'] = "Workshop state days"
	rd1['E1'] = "Workshop national days"
	rd1['F1'] = "Workshop international days"
	rd1['G1'] = "Workshop marks"
	rd1['H1'] = "Presentation state days"
	rd1['I1'] = "Presentation national days"
	rd1['J1'] = "Presentation international days"
	rd1['K1'] = "Presentation marks"
	rd1['L1'] = "No. of online course completed"
	rd1['M1'] = "Online course marks"
	rd1['N1'] = "Sole Authored(No.of Conferences)"
	rd1['O1'] = "First & Second Author(No.of Conferences)"
	rd1['P1'] = "Other Author(No.of Conferences)"
	rd1['Q1'] = "Marks-Sole Authored-Conferences"
	rd1['R1'] = "Marks-First & Second Author-Conferences"
	rd1['S1'] = "Marks-Other Author-Conferences"
	rd1['T1'] = "Sole Authored(No.of Journals)"
	rd1['U1'] = "First & Second Author(No.of Journals)"
	rd1['V1'] = "Other Author(No.of Journals)"
	rd1['W1'] = "Marks-Sole Authored-Journals"
	rd1['X1'] = "Marks-First & Second Author-Journals"
	rd1['Y1'] = "Marks-Other Author-Journals"
	rd1['Z1'] = "Book Authored International"
	rd1['AA1'] = "Book Authored National"
	rd1['AB1'] = "Book chapter Authored International"

	rd1['AC1'] = "Book chapter Authored National"
	rd1['AD1'] = "Article published international"
	rd1['AE1'] = "National Media"
	rd1['AF1'] = "Book Marks"
	rd1['AG1'] = "Internally Funded-Being Sole Principal Investigator"

	rd1['AH1'] = "Internally Funded-Being first Principal Investigator"

	rd1['AI1'] = "Internally Funded-Being Co-Investigator"

	rd1['AJ1'] = "Externally Funded Less than 5_Lakhs-Being Sole Principal Investigator"

	rd1['AK1'] = "Externally Funded Less than 5_Lakhs-Being first Principal Investigator"

	rd1['AL1'] = "Externally Funded Less than 5_Lakhs-Being Co-Investigator"

	rd1['AM1'] = "Externally Funded more than 5_Lakhs-Being Sole Principal Investigator"

	rd1['AN1'] = "Externally Funded more than 5_Lakhs-Being first Principal Investigator"

	rd1['AO1'] = "Externally Funded more than 5_Lakhs-Being Co-Investigator"

	rd1['AP1'] = "Consultancy Work-Upto 2 Lakhs"

	rd1['AQ1'] = "Consultancy Work-Upto 2-5 Lakhs"
	rd1['AR1'] = "Consultancy Work-Above 5 Lakhs"


	rd1['AS1'] = "IPR Info"
	rd1['AT1'] = "Research project marks"
	rd1['AU1'] = "R&D TOTAL MARKS"









	for i,rx in zip(range(2,40),data1):

		info1 = rx.info

		data4 = User.objects.get(username=info1)



		rd1.cell(row = i, column = 1).value = data4.first_name
		rd1.cell(row = i, column = 2).value = data4.department.name
		rd1.cell(row = i, column = 3).value = data4.designation.name







		rd1.cell(row = i, column = 4).value = rx.w_s_d
		rd1.cell(row = i, column = 5).value = rx.w_n_d
		rd1.cell(row = i, column = 6).value = rx.w_i_d
		rd1.cell(row = i, column = 7).value = rx.w_m

		rd1.cell(row = i, column = 8).value = rx.p_s_d
		rd1.cell(row = i, column = 9).value = rx.p_n_d
		rd1.cell(row = i, column = 10).value = rx.p_i_d
		rd1.cell(row = i, column = 11).value = rx.p_m

		rd1.cell(row = i, column = 12).value = rx.onl_course_c
		rd1.cell(row = i, column = 13).value = rx.onl_course_m

		rd1.cell(row = i, column = 14).value = rx.s_c_n
		rd1.cell(row = i, column = 15).value = rx.f_c_n
		rd1.cell(row = i, column = 16).value = rx.o_c_n

		rd1.cell(row = i, column = 17).value = rx.s_c_m
		rd1.cell(row = i, column = 18).value = rx.f_c_m
		rd1.cell(row = i, column = 19).value = rx.o_c_m

		rd1.cell(row = i, column = 20).value = rx.s_j_n
		rd1.cell(row = i, column = 21).value = rx.f_j_n
		rd1.cell(row = i, column = 22).value = rx.o_j_n

		rd1.cell(row = i, column = 23).value = rx.s_j_m
		rd1.cell(row = i, column = 24).value = rx.f_j_m
		rd1.cell(row = i, column = 25).value = rx.o_j_m


		rd1.cell(row = i, column = 26).value = rx.book_i
		rd1.cell(row = i, column = 27).value = rx.book_n
		rd1.cell(row = i, column = 28).value = rx.book_ci
		rd1.cell(row = i, column = 29).value = rx.book_cn
		rd1.cell(row = i, column = 30).value = rx.book_ai
		rd1.cell(row = i, column = 31).value = rx.book_nm


		rd1.cell(row = i, column = 32).value = rx.book_m


		rd1.cell(row = i, column = 33).value = rx.if_s
		rd1.cell(row = i, column = 34).value = rx.if_f
		rd1.cell(row = i, column = 35).value = rx.if_c

		rd1.cell(row = i, column = 36).value = rx.ef_s
		rd1.cell(row = i, column = 37).value = rx.ef_f
		rd1.cell(row = i, column = 38).value = rx.ef_c


		rd1.cell(row = i, column = 39).value = rx.eef_s
		rd1.cell(row = i, column = 40).value = rx.eef_f
		rd1.cell(row = i, column = 41).value = rx.eef_c

		rd1.cell(row = i, column = 42).value = rx.Cw_2
		rd1.cell(row = i, column = 43).value = rx.Cw_2_5
		rd1.cell(row = i, column = 44).value = rx.Cw_5


		rd1.cell(row = i, column = 45).value = rx.ipr_info

		rd1.cell(row = i, column = 46).value = rx.rp_marks
		rd1.cell(row = i, column = 47).value = rx.rd_tot_marks







	journal1['A1'] = "Name"
	journal1['B1'] = "Department"
	journal1['C1'] = "Designation"

	journal1['D1'] = "j1_index"
	journal1['E1'] = "j1_name"
	journal1['F1'] = "j1_title"
	journal1['G1'] = "j1_volume"
	journal1['H1'] = "j1_issn"
	journal1['I1'] = "j1_date"
	journal1['J1'] = "j1_page"
	journal1['K1'] = "j1_author"

	journal1['L1'] = "j2_index"
	journal1['M1'] = "j2_name"
	journal1['N1'] = "j2_title"
	journal1['O1'] = "j2_volume"
	journal1['P1'] = "j2_issn"
	journal1['Q1'] = "j2_date"
	journal1['R1'] = "j2_page"
	journal1['S1'] = "j2_author"


	journal1['T1'] = "j3_index"
	journal1['U1'] = "j3_name"
	journal1['V1'] = "j3_title"
	journal1['W1'] = "j3_volume"
	journal1['X1'] = "j3_issn"
	journal1['Y1'] = "j3_date"
	journal1['Z1'] = "j3_page"
	journal1['AA1'] = "j3_author"



	journal1['AB1'] = "j4_index"
	journal1['AC1'] = "j4_name"
	journal1['AD1'] = "j4_title"
	journal1['AE1'] = "j4_volume"
	journal1['AF1'] = "j4_issn"
	journal1['AG1'] = "j4_date"
	journal1['AH1'] = "j4_page"
	journal1['AI1'] = "j4_author"


	journal1['AJ1'] = "j5_index"
	journal1['AK1'] = "j5_name"
	journal1['AL1'] = "j5_title"
	journal1['AM1'] = "j5_volume"
	journal1['AN1'] = "j5_issn"
	journal1['AO1'] = "j5_date"
	journal1['AP1'] = "j5_page"
	journal1['AQ1'] = "j5_author"



	journal1['AR1'] = "j6_index"
	journal1['AS1'] = "j6_name"
	journal1['AT1'] = "j6_title"
	journal1['AU1'] = "j6_volume"
	journal1['AV1'] = "j6_issn"
	journal1['AW1'] = "j6_date"
	journal1['AX1'] = "j6_page"
	journal1['AY1'] = "j6_author"


	journal1['AZ1'] = "j7_index"
	journal1['BA1'] = "j7_name"
	journal1['BB1'] = "j7_title"
	journal1['BC1'] = "j7_volume"
	journal1['BD1'] = "j7_issn"
	journal1['BE1'] = "j7_date"
	journal1['BF1'] = "j7_page"
	journal1['BG1'] = "j7_author"



	journal1['BH1'] = "j8_index"
	journal1['BI1'] = "j8_name"
	journal1['BJ1'] = "j8_title"
	journal1['BK1'] = "j8_volume"
	journal1['BL1'] = "j8_issn"
	journal1['BM1'] = "j8_date"
	journal1['BN1'] = "j8_page"
	journal1['BO1'] = "j8_author"



	for j,jx in zip(range(2,40),data2):

		info2 = jx.info

		data5 = User.objects.get(username=info2)

		journal1.cell(row = j, column = 1).value = data5.first_name
		journal1.cell(row = j, column = 2).value = data5.department.name
		journal1.cell(row = j, column = 3).value = data5.designation.name






		journal1.cell(row = j, column = 4).value = jx.j1_index
		journal1.cell(row = j, column = 5).value = jx.j1_name
		journal1.cell(row = j, column = 6).value = jx.j1_title
		journal1.cell(row = j, column = 7).value = jx.j1_volume
		journal1.cell(row = j, column = 8).value = jx.j1_issn
		journal1.cell(row = j, column = 9).value = jx.j1_date
		journal1.cell(row = j, column = 10).value = jx.j1_page
		journal1.cell(row = j, column = 11).value = jx.j1_author

		journal1.cell(row = j, column = 12).value = jx.j2_index
		journal1.cell(row = j, column = 13).value = jx.j2_name
		journal1.cell(row = j, column = 14).value = jx.j2_title
		journal1.cell(row = j, column = 15).value = jx.j2_volume
		journal1.cell(row = j, column = 16).value = jx.j2_issn
		journal1.cell(row = j, column = 17).value = jx.j2_date
		journal1.cell(row = j, column = 18).value = jx.j2_page
		journal1.cell(row = j, column = 19).value = jx.j2_author


		journal1.cell(row = j, column = 20).value = jx.j3_index
		journal1.cell(row = j, column = 21).value = jx.j3_name
		journal1.cell(row = j, column = 22).value = jx.j3_title
		journal1.cell(row = j, column = 23).value = jx.j3_volume
		journal1.cell(row = j, column = 24).value = jx.j3_issn
		journal1.cell(row = j, column = 25).value = jx.j3_date
		journal1.cell(row = j, column = 26).value = jx.j3_page
		journal1.cell(row = j, column = 27).value = jx.j3_author


		journal1.cell(row = j, column = 28).value = jx.j4_index
		journal1.cell(row = j, column = 29).value = jx.j4_name
		journal1.cell(row = j, column = 30).value = jx.j4_title
		journal1.cell(row = j, column = 31).value = jx.j4_volume
		journal1.cell(row = j, column = 32).value = jx.j4_issn
		journal1.cell(row = j, column = 33).value = jx.j4_date
		journal1.cell(row = j, column = 34).value = jx.j4_page
		journal1.cell(row = j, column = 35).value = jx.j4_author


		journal1.cell(row = j, column = 36).value = jx.j5_index
		journal1.cell(row = j, column = 37).value = jx.j5_name
		journal1.cell(row = j, column = 38).value = jx.j5_title
		journal1.cell(row = j, column = 39).value = jx.j5_volume
		journal1.cell(row = j, column = 40).value = jx.j5_issn
		journal1.cell(row = j, column = 41).value = jx.j5_date
		journal1.cell(row = j, column = 42).value = jx.j5_page
		journal1.cell(row = j, column = 43).value = jx.j5_author


		journal1.cell(row = j, column = 44).value = jx.j6_index
		journal1.cell(row = j, column = 45).value = jx.j6_name
		journal1.cell(row = j, column = 46).value = jx.j6_title
		journal1.cell(row = j, column = 47).value = jx.j6_volume
		journal1.cell(row = j, column = 48).value = jx.j6_issn
		journal1.cell(row = j, column = 49).value = jx.j6_date
		journal1.cell(row = j, column = 50).value = jx.j6_page
		journal1.cell(row = j, column = 51).value = jx.j6_author

		journal1.cell(row = j, column = 52).value = jx.j7_index
		journal1.cell(row = j, column = 53).value = jx.j7_name
		journal1.cell(row = j, column = 54).value = jx.j7_title
		journal1.cell(row = j, column = 55).value = jx.j7_volume
		journal1.cell(row = j, column = 56).value = jx.j7_issn
		journal1.cell(row = j, column = 57).value = jx.j7_date
		journal1.cell(row = j, column = 58).value = jx.j7_page
		journal1.cell(row = j, column = 59).value = jx.j7_author

		journal1.cell(row = j, column = 60).value = jx.j8_index
		journal1.cell(row = j, column = 61).value = jx.j8_name
		journal1.cell(row = j, column = 62).value = jx.j8_title
		journal1.cell(row = j, column = 63).value = jx.j8_volume
		journal1.cell(row = j, column = 64).value = jx.j8_issn
		journal1.cell(row = j, column = 65).value = jx.j8_date
		journal1.cell(row = j, column = 66).value = jx.j8_page
		journal1.cell(row = j, column = 67).value = jx.j8_author













	conference1['A1'] = "Name"
	conference1['B1'] = "Department"
	conference1['C1'] = "Designation"

	conference1['D1'] = "c1_name"
	conference1['E1'] = "c1_title"
	conference1['F1'] = "c1_place"
	conference1['G1'] = "c1_date"
	conference1['H1'] = "c1_index"
	conference1['I1'] = "c1_author"

	conference1['J1'] = "c2_name"
	conference1['K1'] = "c2_title"
	conference1['L1'] = "c2_place"
	conference1['M1'] = "c2_date"
	conference1['N1'] = "c2_index"
	conference1['O1'] = "c2_author"


	conference1['P1'] = "c3_name"
	conference1['Q1'] = "c3_title"
	conference1['R1'] = "c3_place"
	conference1['S1'] = "c3_date"
	conference1['T1'] = "c3_index"
	conference1['U1'] = "c3_author"


	conference1['V1'] = "c4_name"
	conference1['W1'] = "c4_title"
	conference1['X1'] = "c4_place"
	conference1['Y1'] = "c4_date"
	conference1['Z1'] = "c4_index"
	conference1['AA1'] = "c4_author"

	conference1['AB1'] = "c5_name"
	conference1['AC1'] = "c5_title"
	conference1['AD1'] = "c5_place"
	conference1['AE1'] = "c5_date"
	conference1['AF1'] = "c5_index"
	conference1['AG1'] = "c5_author"


	conference1['AH1'] = "c6_name"
	conference1['AI1'] = "c6_title"
	conference1['AJ1'] = "c6_place"
	conference1['AK1'] = "c6_date"
	conference1['AL1'] = "c6_index"
	conference1['AM1'] = "c6_author"

	conference1['AN1'] = "c7_name"
	conference1['AO1'] = "c7_title"
	conference1['AP1'] = "c7_place"
	conference1['AQ1'] = "c7_date"
	conference1['AR1'] = "c7_index"
	conference1['AS1'] = "c7_author"

	conference1['AT1'] = "c8_name"
	conference1['AU1'] = "c8_title"
	conference1['AV1'] = "c8_place"
	conference1['AW1'] = "c8_date"
	conference1['AX1'] = "c8_index"
	conference1['AY1'] = "c8_author"




	for k,cx in zip(range(2,40),data3):

		info3 = cx.info

		data6 = User.objects.get(username=info3)

		conference1.cell(row = k, column = 1).value = data6.first_name
		conference1.cell(row = k, column = 2).value = data6.department.name
		conference1.cell(row = k, column = 3).value = data6.designation.name






		conference1.cell(row = k, column = 4).value = cx.c1_index
		conference1.cell(row = k, column = 5).value = cx.c1_name
		conference1.cell(row = k, column = 6).value = cx.c1_place
		conference1.cell(row = k, column = 7).value = cx.c1_date
		conference1.cell(row = k, column = 8).value = cx.c1_index
		conference1.cell(row = k, column = 9).value = cx.c1_author


		conference1.cell(row = k, column = 10).value = cx.c2_index
		conference1.cell(row = k, column = 11).value = cx.c2_name
		conference1.cell(row = k, column = 12).value = cx.c2_place
		conference1.cell(row = k, column = 13).value = cx.c2_date
		conference1.cell(row = k, column = 14).value = cx.c2_index
		conference1.cell(row = k, column = 15).value = cx.c2_author


		conference1.cell(row = k, column = 16).value = cx.c3_index
		conference1.cell(row = k, column = 17).value = cx.c3_name
		conference1.cell(row = k, column = 18).value = cx.c3_place
		conference1.cell(row = k, column = 19).value = cx.c3_date
		conference1.cell(row = k, column = 20).value = cx.c3_index
		conference1.cell(row = k, column = 21).value = cx.c3_author


		conference1.cell(row = k, column = 22).value = cx.c4_index
		conference1.cell(row = k, column = 23).value = cx.c4_name
		conference1.cell(row = k, column = 24).value = cx.c4_place
		conference1.cell(row = k, column = 25).value = cx.c4_date
		conference1.cell(row = k, column = 26).value = cx.c4_index
		conference1.cell(row = k, column = 27).value = cx.c4_author


		conference1.cell(row = k, column = 28).value = cx.c5_index
		conference1.cell(row = k, column = 29).value = cx.c5_name
		conference1.cell(row = k, column = 30).value = cx.c5_place
		conference1.cell(row = k, column = 31).value = cx.c5_date
		conference1.cell(row = k, column = 32).value = cx.c5_index
		conference1.cell(row = k, column = 33).value = cx.c5_author

		conference1.cell(row = k, column = 34).value = cx.c6_index
		conference1.cell(row = k, column = 35).value = cx.c6_name
		conference1.cell(row = k, column = 36).value = cx.c6_place
		conference1.cell(row = k, column = 37).value = cx.c6_date
		conference1.cell(row = k, column = 38).value = cx.c6_index
		conference1.cell(row = k, column = 39).value = cx.c6_author


		conference1.cell(row = k, column = 40).value = cx.c7_index
		conference1.cell(row = k, column = 41).value = cx.c7_name
		conference1.cell(row = k, column = 42).value = cx.c7_place
		conference1.cell(row = k, column = 43).value = cx.c7_date
		conference1.cell(row = k, column = 44).value = cx.c7_index
		conference1.cell(row = k, column = 45).value = cx.c7_author


		conference1.cell(row = k, column = 46).value = cx.c8_index
		conference1.cell(row = k, column = 47).value = cx.c8_name
		conference1.cell(row = k, column = 48).value = cx.c8_place
		conference1.cell(row = k, column = 49).value = cx.c8_date
		conference1.cell(row = k, column = 50).value = cx.c8_index
		conference1.cell(row = k, column = 51).value = cx.c8_author


	response = HttpResponse(save_virtual_workbook(book), content_type='application/vnd.ms-excel')
	response['Content-Disposition'] = 'attachment; filename="final_report.xlsx"'
	return response

def ao_consolidated(request,y):
	data1 = User.objects.filter(Q(department__name='CSE')|Q(department__name='ISE')|Q(department__name='ECE')|Q(department__name='EEE')|Q(department__name='CIV')|Q(department__name='MCA')|Q(department__name='TCE')|Q(department__name='MECH')|Q(department__name='maths')|Q(department__name='physics')|Q(department__name='chemistry'))
	dic = {}

	for i in data1:

		try:
		    data2 = feedbackTab.objects.filter(info=i.pk).get(year=y)
		except feedbackTab.DoesNotExist:
		    data2 = None

		try:
		    data3 = rd.objects.filter(info=i.pk).get(year=y)
		except rd.DoesNotExist:
		    data3 = None

		try:
		    data4 = remarks1.objects.filter(info=i.pk).get(year=y)
		except remarks1.DoesNotExist:
		    data4 = None

		try:
		    data5 = remarks2.objects.filter(info=i.pk).get(year=y)
		except remarks2.DoesNotExist:
		    data5 = None

		try:
		    data7 = empDetailForm.objects.filter(info=i.pk).get(year=y)
		except empDetailForm.DoesNotExist:
		    data7 = None

		dic[i] = [data2,data3,data4,data5,data7]

	context = {
			'data6':dic
		}

	return render(request,'ao_consolidated.html',context = context)

def email_otp(random_otp,email,name):
	mail_subject = 'Your OTP is :.'

	body="Hi ,"\
		+name\
		+"\n\n"\
		+"Your OTP is : "\
		+random_otp\
		+"\n\nThanks,\nPBAS-BMSIT"
	message = str(MIMEText(body,'plain'))

	#print(rand)
	#subject='Thank you for registering with us'
	#message='Please enter the OTP to verify your email.Your OTP is : '+str(rand)
	#message=str(rand)
	from_email='ravikrsngh.rks@gmail.com'
	to_list=[email]
	print(to_list)
	s= smtplib.SMTP('smtp.gmail.com',587)
	s.starttls()
	s.login('ravikrsngh.rks@gmail.com','ykigquucffuhzmkz')
	s.sendmail(from_email,to_list,message)
	s.quit()
	print("Email Sent")



def phone_otp(random_otp, phone):
		phone1 = str(phone)
		message = 'Please login with the OTP: '+random_otp
		params = { 'number' : phone1, 'text' : message }
		baseUrl = 'https://www.smsgatewayhub.com/api/mt/SendSMS?APIKey=62sxGWT6MkCjDul6eNKejw&senderid=BMSITM&channel=2&DCS=0&flashsms=0&' + ap.urlencode(params)
		print("PHONE OTP SENT")
		urllib.request.urlopen(baseUrl).read(1000)


def invalid(request):
	return render(request,'invalid.html')

def login(request):
	if(request.method=='POST'):

		userna = request.POST.get('username')
		username = userna.upper()

		if(User.objects.filter(username=username).exists()):
			user = User.objects.get(username=username)
			first_name = user.first_name
			email = user.email
			phone = user.phone
			random_otp = r''.join(random.choice('0123456789') for i in range(4))
			phone_otp(random_otp,phone)

			email_otp(random_otp,user.email,user.first_name)
			#random_otp="1234"
			hashed_pwd = make_password(random_otp)
			User.objects.filter(username=username).update(password=hashed_pwd)

			return HttpResponseRedirect("/login/user=" + username)

	return render(request,'login.html')

def decide_view(request):
	user=request.user
	if user.designation.pk is 5 or  user.designation.pk is 6:
		return HttpResponseRedirect(reverse('facform1:ao_principal'))
	return HttpResponseRedirect(reverse('facform1:first_page'))
	"""
	if request.user.is_assistant_professor():
		if request.user.teach_status == True:
			return HttpResponseRedirect(reverse('facform1:first_page'))
		return HttpResponseRedirect(reverse('facform1:first_page'))

	elif request.user.is_associate_professor():
		if request.user.teach_status == True:
			return HttpResponseRedirect("/associate_preview/")
		return HttpResponseRedirect(reverse('facform1:first_page'))

	elif request.user.is_professor():
		if request.user.teach_status == True:
			return HttpResponseRedirect("/associate_preview/")
		return HttpResponseRedirect(reverse('facform1:first_page'))

	elif request.user.is_hod():
		# print('hod')
		return HttpResponseRedirect("/hod_first/")

	elif request.user.is_principal():
		print('princy')
		return HttpResponseRedirect("/principal_first/")

	elif request.user.is_ao():
		return HttpResponseRedirect("/ao_first/")
"""
	# else:
	# 	return HttpResponseRedirect("/")
# def front(request):
# 	return render(request,'front.html')






@login_required
def hod_display(request,y):
	user = request.user
	hod_dept = user.department
	if request.user.is_hod() and request.user.department == hod_dept:

		user = request.user
		pk = request.user.pk

		hod_desg = user.designation
		c = User.objects.filter(department=hod_dept)
		print(c)
		c1=set()
		c2=set()
		c3=set()
		for i in c:
			if i.designation.pk is 1:
				try:
					x=new.objects.filter(info__username=i.username).get(year__year=y)
					print(x.teach_status)
					c1.add(x)
				except:
					pass
			if i.designation.pk is 2:
				try:
					x=new.objects.filter(info__username=i.username).get(year__year=y)
					print(x.teach_status)
					c2.add(x)
				except:
					pass

			if i.designation.pk is 3:
				try:
					x=new.objects.filter(info__username=i.username).get(year__year=y)
					print(x.teach_status)
					c3.add(x)
				except:
					pass
		print(c1)
		print(c2)
		print(c3)
		# c2 = .objects.filter(info__department=hod_dept).filter(~Q(info__designation=hod_desg))

		context = {'name':c1,'name1':c2 ,'hod_dept':hod_dept,'y':y,'name2':c3}

		# print(data6)
		return render(request,'hod_display.html',context=context)
	else:
		return HttpResponseRedirect('/invalid')

@login_required
def hod_teacher_display(request,pk,y):
		if request.user.is_hod():
			name =  User.objects.get(pk = pk);
			print(name)
			data1 = User.objects.get(username=name);
			print(data1.first_name)
			data2 = empDetailForm.objects.filter(info__username=name).get(year=y);
			data3 = feedbackTab.objects.filter(info__username=name).get(year=y);
			data4 = rd.objects.filter(info__username=name).get(year=y);
			data5 = remarks.objects.filter(info__username=name).get(year=y);
			data6 = conference.objects.filter(info__username=name).get(year=y);
			data7 = journal.objects.filter(info__username=name).get(year=y);






			if request.method == 'POST':
				form1 = forms.form_remarks1(request.POST)
				if form1.is_valid():

					sendme = new.objects.filter(info__username=name).get(year__year=y)
					obj = form1.save(commit=False)

					obj.info = name
					obj.year=y
					obj.save()

					if sendme.hod_status == False:
						sendme.hod_status = True
						sendme.save()
					return HttpResponseRedirect(reverse('facform1:hod_display',args=(y,)))

			else:
				form1 = forms.form_remarks1()


			context1 = {
			"key1":data1,
			"key2":data2,
			"key3":data3,
			"key4":data4,
			"key5":data5,

			"key6":data6,
			"key7":data7,
			"form1":form1
			}

			return render(request,'hod_teacher_display.html',context=context1)
		else:
			return HttpResponseRedirect('/invalid')



@login_required
def hod_teacher_display_edit(request,pk,y):
	if request.user.is_hod():
		name =  User.objects.get(pk = pk);
		print(name)
		data1 = User.objects.get(username=name);
		print(data1.first_name)
		data2 = empDetailForm.objects.filter(info__username=name).get(year=y);
		data3 = feedbackTab.objects.filter(info__username=name).get(year=y);
		data4 = rd.objects.filter(info__username=name).get(year=y);
		data5 = remarks.objects.filter(info__username=name).get(year=y);
		data6 = conference.objects.filter(info__username=name).get(year=y);
		data7 = journal.objects.filter(info__username=name).get(year=y);

		# data8 = remarks1.objects.get(info__username=name);
		# data9 = remarks2.objects.get(info__username=name);

		if new.objects.filter(info=name).filter(hod_status=True).filter(year__year=y):
			data8 = remarks1.objects.get(info__username=name)
		else:
			data8 = []
		if new.objects.filter(info=name).filter(year__year=y).filter(principal_status=True):
			data9 = remarks2.objects.get(info__username=name)
		else:
			data9 = []

		if request.method == 'POST':
			form1 = forms.form_remarks1(request.POST,instance=data8)
			if form1.is_valid():

				sendme = new.objects.filter(info__username=name).get(year__year=y)
				obj = form1.save(commit=False)

				obj.info = name
				obj.year=y
				obj.save()

				if sendme.hod_status == False:
					sendme.hod_status = True
					sendme.save()
				return HttpResponseRedirect(reverse('facform1:hod_display',args=(y,)))

		else:
			form1 = forms.form_remarks1(instance=data8)




		# if request.method == 'POST':
		# 	remark  = remarks1.objects.get(info=name)
		# 	form1 = forms.form_remarks1(request.POST, instance=remark)
		# 	if form1.is_valid():

		# 		obj = form1.save(commit=False)
		# 		obj.save()
		# 		return HttpResponseRedirect("/hod_first/")

		# else:
		# 	form1 = forms.form_remarks1()


		context1 = {
		"key1":data1,
		"key2":data2,
		"key3":data3,
		"key4":data4,
		"key5":data5,
		"key6":data6,
		"key7":data7,
		"key8":form1,
		"key9":data9,
		# "form1":form1
		}

		return render(request,'hod_teacher_display_edit.html',context=context1)
	else:
		return HttpResponseRedirect('/invalid')



@login_required
def hod_teacher1_display(request,pk,y):
	if request.user.is_hod():
		name =  User.objects.get(pk = pk);
		print(name)
		data1 = User.objects.get(username=name);
		print(data1.first_name)
		data2 = empDetailForm.objects.filter(info__username=name).get(year=y);
		data3 = feedbackTab.objects.filter(info__username=name).get(year=y);
		data4 = rd.objects.filter(info__username=name).get(year=y);
		data5 = remarks.objects.filter(info__username=name).get(year=y);
		data6 = conference.objects.filter(info__username=name).get(year=y);
		data7 = journal.objects.filter(info__username=name).get(year=y);





		if request.method == 'POST':
			form1 = forms.form_remarks1(request.POST)
			if form1.is_valid():

				sendme = new.objects.filter(username=name).get(year=y)
				obj = form1.save(commit=False)

				obj.info = name
				obj.year=y
				obj.save()

				if sendme.hod_status == False:
					sendme.hod_status = True
					sendme.save()
				return HttpResponseRedirect(reverse('facform1:hod_first',args=(y,)))

		else:
			form1 = forms.form_remarks1()


		context1 = {
		"key1":data1,
		"key2":data2,
		"key3":data3,
		"key4":data4,
		"key5":data5,
		"key6":data6,
		"key7":data7,
		"form1":form1,
		"y":y
		}

		return render(request,'hod_teacher1_display.html',context=context1)

	else:
		return HttpResponseRedirect('/invalid')




@login_required
def hod_teacher1_display_edit(request,pk):
		if request.user.is_hod():
			name =  User.objects.get(pk = pk);
			print(name)
			data1 = User.objects.get(username=name);
			print(data1.first_name)
			data2 = empDetailForm.objects.get(info__username=name);
			data3 = feedbackTab.objects.get(info__username=name);
			data4 = rd.objects.get(info__username=name);
			data5 = remarks.objects.get(info__username=name);
			data6 = conference.objects.get(info__username=name);
			data7 = journal.objects.get(info__username=name);
			# data8 = remarks1.objects.get(info__username=name);
			if User.objects.filter(username=name).filter(hod_status=True):
				data8 = remarks1.objects.get(info__username=name)
			else:
				data8 = []
			if User.objects.filter(username=name).filter(principal_status=True):
				data9 = remarks2.objects.get(info__username=name)
			else:
				data9 = []






			# if request.method == 'POST':
			# 	remark = remarks1.objects.get(info=name)
			# 	form1 = forms.form_remarks1(request.POST, instance=remark)
			# 	if form1.is_valid():

			# 		obj = form1.save(commit=False)
			# 		obj.save()
			# 		return HttpResponseRedirect("/hod_first/")

			# else:
			# 	form1 = forms.form_remarks1()


			context1 = {
			"key1":data1,
			"key2":data2,
			"key3":data3,
			"key4":data4,
			"key5":data5,
			"key6":data6,
			"key7":data7,
			"key8":data8,
			"key9":data9,

			# "form1":form1
			}

			return render(request,'hod_teacher1_display_edit.html',context=context1)
		else:
			return HttpResponseRedirect('/invalid')




@login_required
def principal_first(request,y):
	if request.user.is_principal():
		dept = Department.objects.all()

		context = {'dept':dept,'y':y}


		return render(request,'principal_first.html',context=context)
	else:
		return HttpResponseRedirect('/invalid')


@login_required
def principal_display(request,dept,y):
	if request.user.is_principal():
		user=User.objects.filter(department__name=dept)
		print(user)
		print(dept)
		hod=set()
		teach3=set()
		teach2=set()
		for x in user:
			print(x)
			print(x.designation.pk)
			if x.designation.pk == 1:
				try:
					i=new.objects.filter(info__username=x.username).get(year__year=y)
					print(i.teach_status)
					teach2.add(i)
					print(i)
				except:
					pass
			if x.designation.pk is 2:
				try:
					i=new.objects.filter(info__username=x.username).get(year__year=y)
					print(i.teach_status)
					teach3.add(i)
					print(i)
				except:
					pass
			if x.designation.pk is 3:
				try:
					i=new.objects.filter(info__username=x.username).get(year__year=y)
					print(i.teach_status)
					print(i)
					teach3.add(i)
				except:
					pass
			if x.designation.pk is 4:
				try:
					i=new.objects.filter(info__username=x.username).get(year__year=y)
					print(i.teach_status)
					print(i)
					hod.add(i)
				except:
					pass

		print(teach2)
		print(teach3)
		print(hod)
		return render(request,'principal_display.html',{'dept':dept,'hod':hod,'teach3':teach3,'teach2':teach2,'y':y})
	else:
			return HttpResponseRedirect('/invalid')
"""
		hod = User.objects.filter(department__name=dept).filter(designation__pk = 8).filter(teach_status=True)

		teach = User.objects.filter(department__name=dept).filter(designation__pk = 9).filter(hod_status=True)
		teach1 = User.objects.filter(department__name=dept).filter(designation__pk = 10).filter(hod_status=True)

		teach3 = list(chain(teach,teach1))
		teach2 = User.objects.filter(department__name=dept).filter(designation__pk = 11).filter(hod_status=True)
"""



@login_required
def principal_teacher_display(request,pk,y):
	if request.user.is_principal():
		name =  User.objects.get(pk = pk);
		print(name)
		data1 = User.objects.get(username=name);
		print(data1.department)

		data2 = empDetailForm.objects.filter(info__username=name).get(year=y);
		data3 = feedbackTab.objects.filter(info__username=name).get(year=y);
		data4 = rd.objects.filter(info__username=name).get(year=y);
		data5 = remarks.objects.filter(info__username=name).get(year=y);
		data6 = conference.objects.filter(info__username=name).get(year=y);
		data7 = journal.objects.filter(info__username=name).get(year=y);
		data8 = remarks1.objects.filter(info__username=name).get(year=y);




		if request.method == 'POST':
			form1 = forms.form_remarks2(request.POST)

			if form1.is_valid():
				print("logout")
				sendme = User.objects.get(username=name)
				obj = form1.save(commit=False)

				obj.info = name
				obj.year=y
				obj.department = data1.department
				obj.save()

				sendme=new.objects.filter(info__username=name).get(year__year=y)
				if sendme.principal_status == False:
					sendme.principal_status = True
					sendme.save()

				return HttpResponseRedirect(reverse('facform1:principal_display',args=(data1.department.name,y)))

		else:
			form1 = forms.form_remarks2()


		context1 = {
		"key1":data1,
		"key2":data2,
		"key3":data3,
		"key4":data4,
		"key5":data5,
		"key6":data6,
		"key7":data7,
		"key8":data8,
		"form1":form1,
		"y":y
		}

		return render(request,'principal_teacher_display.html',context=context1)
	else:
		return HttpResponseRedirect('/invalid')


@login_required
def principal_teacher_display_edit(request,pk):
	if request.user.is_principal():
		name =  User.objects.get(pk = pk);
		print(name)
		data1 = User.objects.get(username=name);
		print(data1.department)

		data2 = empDetailForm.objects.get(info__username=name);
		data3 = feedbackTab.objects.get(info__username=name);
		data4 = rd.objects.get(info__username=name);
		data5 = remarks.objects.get(info__username=name);
		data6 = conference.objects.get(info__username=name);
		data7 = journal.objects.get(info__username=name);
		data8 = remarks1.objects.get(info__username=name);
		data9 = remarks2.objects.get(info__username=name);




		if request.method == 'POST':
			remark  = remarks2.objects.get(info=name)
			form1 = forms.form_remarks2(request.POST, instance=remark)

			if form1.is_valid():
				obj = form1.save(commit=False)
				obj.info = name
				obj.department = data1.department
				obj.save()


				return HttpResponseRedirect("/principal_display/"+ data1.department.name)

		else:
			form1 = forms.form_remarks2()


		context1 = {
		"key1":data1,
		"key2":data2,
		"key3":data3,
		"key4":data4,
		"key5":data5,
		"key6":data6,
		"key7":data7,
		"key8":data8,
		"key9":data9,
		"form1":form1
		}

		return render(request,'principal_teacher_display_edit.html',context=context1)
	else:
		return HttpResponseRedirect('/invalid')


@login_required
def principal_teacher1_display(request,pk,y):
	if request.user.is_principal():
		name =  User.objects.get(pk = pk);
		print(name)
		data1 = User.objects.get(username=name);
		print(data1.department)

		data2 = empDetailForm.objects.filter(info__username=name).get(year=y);
		data3 = feedbackTab.objects.filter(info__username=name).get(year=y);
		data4 = rd.objects.filter(info__username=name).get(year=y);
		data5 = remarks.objects.filter(info__username=name).get(year=y);
		data6 = conference.objects.filter(info__username=name).get(year=y);
		data7 = journal.objects.filter(info__username=name).get(year=y);
		if new.objects.filter(info__username=name).filter(hod_status=True):
			data8 = remarks1.objects.filter(info__username=name).get(year=y)
		else:
			data8 = []



		if request.method == 'POST':
			form1 = forms.form_remarks2(request.POST)

			if form1.is_valid():
				print("logout")
				#sendme = User.objects.get(username=name)
				obj = form1.save(commit=False)

				obj.info = name
				obj.year=y
				obj.department = data1.department
				obj.save()
				sendme=new.objects.filter(info__username=name).get(year__year=y)
				if sendme.principal_status == False:
					sendme.principal_status = True
					sendme.save()

				return HttpResponseRedirect(reverse('facform1:principal_display',args=(data1.department.name,y,)))

		else:
			form1 = forms.form_remarks2()


		context1 = {
		"key1":data1,
		"key2":data2,
		"key3":data3,
		"key4":data4,
		"key5":data5,
		"key6":data6,
		"key7":data7,
		"key8":data8,
		"form1":form1,
		"y":y
		}

		return render(request,'principal_teacher1_display.html',context=context1)
	else:
		return HttpResponseRedirect('/invalid')


@login_required
def principal_teacher1_display_edit(request,pk):
	if request.user.is_principal():
		name =  User.objects.get(pk = pk);
		print(name)
		data1 = User.objects.get(username=name);
		print(data1.department)

		data2 = empDetailForm.objects.get(info__username=name);
		data3 = feedbackTab.objects.get(info__username=name);
		data4 = rd.objects.get(info__username=name);
		data5 = remarks.objects.get(info__username=name);
		data6 = conference.objects.get(info__username=name);
		data7 = journal.objects.get(info__username=name);
		if User.objects.filter(username=name).filter(hod_status=True):
			data8 = remarks1.objects.get(info__username=name)
		else:
			data8 = []

		if User.objects.filter(username=name).filter(principal_status=True):
			data9 = remarks2.objects.get(info__username=name);
		else:
			data9 = []



		if request.method == 'POST':
			remark  = remarks2.objects.get(info=name)
			form1 = forms.form_remarks2(request.POST, instance=remark)

			if form1.is_valid():
				obj = form1.save(commit=False)

				obj.info = name
				obj.department = data1.department
				obj.save()

				return HttpResponseRedirect("/principal_display/"+data1.department.name)

		else:
			form1 = forms.form_remarks2()


		context1 = {
		"key1":data1,
		"key2":data2,
		"key3":data3,
		"key4":data4,
		"key5":data5,
		"key6":data6,
		"key7":data7,
		"key8":data8,
		"key9":data9,
		"form1":form1
		}

		return render(request,'principal_teacher1_display_edit.html',context=context1)
	else:
		return HttpResponseRedirect('/invalid')




@login_required
def principal_hod_display(request,pk,y):
	if request.user.is_principal():
		name =  User.objects.get(pk = pk);
		print(name)
		data1 = User.objects.get(username=name);
		print(data1.department)

		data2 = empDetailForm.objects.filter(info__username=name).get(year=y);
		data3 = feedbackTab.objects.filter(info__username=name).get(year=y);
		data4 = rd.objects.filter(info__username=name).get(year=y);
		data5 = remarks.objects.filter(info__username=name).get(year=y);
		data6 = conference.objects.filter(info__username=name).get(year=y);
		data7 = journal.objects.filter(info__username=name).get(year=y);




		if request.method == 'POST':
			form1 = forms.form_remarks2(request.POST)

			if form1.is_valid():
				print("logout")
				sendme = User.objects.get(username=name)
				obj = form1.save(commit=False)

				obj.info = name
				obj.year=y
				obj.department = data1.department
				obj.save()
				sendme=new.objects.filter(info__username=name).get(year__year=y)
				if sendme.principal_status == False:
					sendme.principal_status = True
					sendme.save()

				return HttpResponseRedirect(reverse('facform1:principal_display',args=(data1.department.name,y,)))

		else:
			form1 = forms.form_remarks2()


		context1 = {
		"key1":data1,
		"key2":data2,
		"key3":data3,
		"key4":data4,
		"key5":data5,
		"key6":data6,
		"key7":data7,
		"form1":form1,
		"y":y,
		}

		return render(request,'principal_hod_display.html',context=context1)
	else:
		return HttpResponseRedirect('/invalid')



@login_required
def principal_hod_display_edit(request,pk):
	if request.user.is_principal():
		name =  User.objects.get(pk = pk);
		print(name)
		data1 = User.objects.get(username=name);
		print(data1.department)

		data2 = empDetailForm.objects.get(info__username=name);
		data3 = feedbackTab.objects.get(info__username=name);
		data4 = rd.objects.get(info__username=name);
		data5 = remarks.objects.get(info__username=name);
		data6 = conference.objects.get(info__username=name);
		data7 = journal.objects.get(info__username=name);
		if User.objects.filter(username=name).filter(hod_status=True):
			data8 = remarks2.objects.get(info__username=name)
		else:
			data8 = []




		if request.method == 'POST':
			remark  = remarks2.objects.get(info=name)
			form1 = forms.form_remarks2(request.POST, instance=remark)

			if form1.is_valid():
				obj = form1.save(commit=False)

				obj.info = name
				obj.department = data1.department
				obj.save()

				return HttpResponseRedirect("/principal_display/"+data1.department.name)

		else:
			form1 = forms.form_remarks2()


		context1 = {
		"key1":data1,
		"key2":data2,
		"key3":data3,
		"key4":data4,
		"key5":data5,
		"key6":data6,
		"key7":data7,
		"key8":data8,
		"form1":form1
		}

		return render(request,'principal_hod_display_edit.html',context=context1)
	else:
		return HttpResponseRedirect('/invalid')


@login_required
def ao_first(request,y):
	if request.user.is_ao():
		dept = Department.objects.all()
		context = {'dept':dept , 'y':y}
		print(y)
		return render(request,'ao_first.html',context=context)
	else:
		return HttpResponseRedirect('/invalid')



@login_required
def ao_display(request,dept,y):
	if request.user.is_ao():

		print(dept)
		a = "assistant"
		b = "associate"
		c = "professor"
		d = "hod"
		# above = remarks2.objects.filter(total_marks__gte = 60).filter(department__name=dept).filter(info__designation__pk=11)
		# above1 = remarks2.objects.filter(total_marks__gte = 60).filter(department__name=dept).filter(info__designation__pk=9)
		# above2 = remarks2.objects.filter(total_marks__gte = 60).filter(department__name=dept).filter(info__designation__pk=10)
		# above3 = remarks2.objects.filter(total_marks__gte = 60).filter(department__name=dept).filter(info__designation__pk=8)
		#
		# above4 = list(chain(above1,above2))
		#
		# below = remarks2.objects.filter(total_marks__lt = 60).filter(department__name=dept).filter(info__designation__pk=11)
		# below1 = remarks2.objects.filter(total_marks__lt = 60).filter(department__name=dept).filter(info__designation__pk=9)
		# below2 = remarks2.objects.filter(total_marks__lt = 60).filter(department__name=dept).filter(info__designation__pk=10)
		# below3 = remarks2.objects.filter(total_marks__lt = 60).filter(department__name=dept).filter(info__designation__pk=8)
		#
		# below4 = list(chain(below1,below2))
		general = User.objects.filter(department__name=dept)

		tt = general.count()
		teach_completed=0
		hod_completed=0
		principal_completed=0
		#sub=new.objects.filter(year__year=y)
		sub=set()
		for i in general:
			try:
				s=new.objects.filter(info__username=i).get(year__year=y)
				print(s)
				sub.add(s);
				if s.teach_status:
					teach_completed=teach_completed+1
				if s.hod_status:
					hod_completed=hod_completed+1
				if s.principal_status:
					principal_completed=principal_completed+1
			except:
				pass











		context = {'dept':dept,'general':general,'p':principal_completed,'h':hod_completed,'t':teach_completed,'total':tt,'sub':sub,'y':y}

		return render(request,'ao_display.html',context=context)
	else:
		return HttpResponseRedirect('/invalid')

@login_required
def ao_approved(request,dept,y):
	if request.user.is_ao():


		general = remarks2.objects.filter(department__name=dept).filter(year=y)
		dept = {'dept':dept,'general':general,'y':y}
		return render(request,'ao_approved.html',context=dept)
	else:
		return HttpResponseRedirect('/invalid')



@login_required
def ao_teacher_display(request,name,y):
	print(name)
	data1 = User.objects.get(username=name)
	data2 = empDetailForm.objects.filter(info__username=name).get(year=y)
	data3 = feedbackTab.objects.filter(info__username=name).get(year=y)
	data4 = rd.objects.filter(info__username=name).get(year=y)
	data5 = remarks.objects.filter(info__username=name).get(year=y)
	data6 = conference.objects.filter(info__username=name).get(year=y)
	data7 = journal.objects.filter(info__username=name).get(year=y)

	if new.objects.filter(info__username=name).filter(hod_status=True):
		data8 = remarks1.objects.filter(info__username=name).get(year=y)
	else:
		data8 = []
	if new.objects.filter(info__username=name).filter(principal_status=True):
		data9 = remarks2.objects.filter(info__username=name).get(year=y)
	else:
		data9 = []
	# if remarks1.objects.get(info__username=name):
	# 	data8 = remarks1.objects.get(info__username=name)
	#
	# if remarks2.objects.get(info__username=name):
	# 	data9 = remarks2.objects.get(info__username=name)







	context1 = {
	"key1":data1,
	"key2":data2,
	"key3":data3,
	"key4":data4,
	"key5":data5,
	"key6":data6,
	"key7":data7,
	"key8":data8,
	"key9":data9,


	}




	# if request.method == 'POST':
	# 	email_doc(name)
	# 	return HttpResponseRedirect('/ao_display'+ data1.department)

	return render(request,'ao_teacher_display.html',context=context1)

@login_required
def ao_teacher1_display(request,name,y):
	print(name)
	data1 = User.objects.get(username=name)
	data2 = empDetailForm.objects.filter(info__username=name).get(year=y)
	data3 = feedbackTab.objects.filter(info__username=name).get(year=y)
	data4 = rd.objects.filter(info__username=name).get(year=y)
	data5 = remarks.objects.filter(info__username=name).get(year=y)
	data6 = conference.objects.filter(info__username=name).get(year=y)
	data7 = journal.objects.filter(info__username=name).get(year=y)

	if new.objects.filter(info__username=name).filter(hod_status=True):
		data8 = remarks1.objects.filter(info__username=name).get(year=y)
	else:
		data8 = []
	if new.objects.filter(info__username=name).filter(principal_status=True):
		data9 = remarks2.objects.filter(info__username=name).get(year=y)
	else:
		data9 = []



	context1 = {
	"key1":data1,
	"key2":data2,
	"key3":data3,
	"key4":data4,
	"key5":data5,
	"key6":data6,
	"key7":data7,
	"key8":data8,
	"key9":data9,


	}
	return render(request,'ao_teacher1_display.html',context=context1)

@login_required
def ao_hod_display(request,name,y):
	if request.user.is_ao():
		print(name)
		data1 = User.objects.get(username=name)
		data2 = empDetailForm.objects.filter(info__username=name).get(year=y)
		data3 = feedbackTab.objects.filter(info__username=name).get(year=y)
		data4 = rd.objects.filter(info__username=name).get(year=y)
		data5 = remarks.objects.filter(info__username=name).get(year=y)
		data6 = conference.objects.filter(info__username=name).get(year=y)
		data7 = journal.objects.filter(info__username=name).get(year=y)
		if new.objects.filter(info__username=name).filter(principal_status=True):
			data8 = remarks2.objects.filter(info__username=name).get(year=y)
		else:
			data8 = []



		context1 = {
		"key1":data1,
		"key2":data2,
		"key3":data3,
		"key4":data4,
		"key5":data5,
		"key6":data5,
		"key7":data7,
		"key8":data8,


		}
		return render(request,'ao_hod_display.html',context=context1)
	else:
		return HttpResponseRedirect('/invalid')

@login_required
def hod_first(request,y):
	if request.user.is_hod():
		user = request.user
		hod_dept = user.department
		status=new.objects.filter(info=user).get(year__year=y)
		print(status)
		context = {'dept':hod_dept,'user':user,'f':y,'t':status}
		return render(request,'hod_first.html',context=context)

	else:
		print("HI")
		return HttpResponseRedirect('/invalid')




def logout(request):
	return render(request,'hod_success.html')

@login_required
def f_assistant5(request,y):
	user=User.objects.get(username=request.user)
	print(user.designation)
	if user.is_assistant_professor():

		form6 = forms.form_conference()
		form7 = forms.form_journal()

		if conference.objects.filter(info=request.user).filter(year=y).exists():

			return HttpResponseRedirect(reverse('facform1:assistant_preview',args=(y,)))
		else:
			if request.method == 'POST':
				sendme = User.objects.get(username=request.user)
				form6 = forms.form_conference(request.POST,request.FILES)
				form7 = forms.form_journal(request.POST,request.FILES)
				if form6.is_valid() and form7.is_valid():
					obj3 = form6.save(commit=False)
					obj4 = form7.save(commit=False)
					obj3.info = request.user
					obj4.info = request.user

					obj3.year=y
					obj4.year=y

					obj3.save()
					obj4.save()
					return HttpResponseRedirect(reverse('facform1:f_assistant5_final',args=(y,)))
			else:
				print(form6.errors)
				print(form7.errors)
			return render(request,'assistant_form5.html',{'form6':form6,'form7':form7,'y':y})
		return render(request,'assistant_form5.html',{'form6':form6,'form7':form7,'y':y})
	else:
		return HttpResponseRedirect('/invalid')

@login_required
def f_assistant_edit5(request,y):
	user=request.user
	obj1=conference.objects.filter(info=user).get(year=y)
	obj2=journal.objects.filter(info=user).get(year=y)
	form2= forms.form_conference(instance=obj1)
	form3=forms.form_journal(instance=obj2)
	if request.method == 'POST':
		# form1 = forms.form_User(request.POST)
		#form3 = form3(request.POST,request.FILES)
		form2 = forms.form_conference(request.POST,request.FILES,instance=obj1)
		form3=forms.form_journal(request.POST,request.FILES,instance=obj2)

		if form2.is_valid() and form3.is_valid():

			# sendme = User.objects.get(username=request.user)
			obj1 = form2.save(commit=False)
			obj1.info = request.user
			obj1.year=y
			obj1.save()
			obj2 = form3.save(commit=False)
			obj2.info = request.user
			obj2.year=y
			obj2.save()
			# sendme.doc_link  = 	form1.cleaned_data['doc_link']
			# sendme.save()

			return HttpResponseRedirect(reverse('facform1:f_assistant5_final',args=(y,)))
	return render(request,'assistant_form5.html',{'form6':form2,'form7':form3,'y':y})

@login_required
def status(request,y):
	sendme = User.objects.get(username=request.user)
	n=new.objects.filter(info=request.user).get(year__year=y)
	if n.teach_status == False:
		n.teach_status = True
		n.save()
	return HttpResponseRedirect(reverse('facform1:logout'))


	return HttpResponseRedirect(reverse('facform1:logout'))
@login_required
def f_assistant5_final(request,y):
	user=request.user

	key6=conference.objects.filter(info=user).get(year=y)
	key7=journal.objects.filter(info=user).get(year=y)
	return render(request,'f_assistant5_final.html',{'key6':key6,'key7':key7,'y':y})


@login_required
def f_assistant4(request,y):
	if request.user.is_assistant_professor():
		form5 = forms.form_remarks()

		if remarks.objects.filter(info=request.user).filter(year=y).exists():

			return HttpResponseRedirect(reverse('facform1:assistant_form5',args=(y,)))
		else:
			if request.method == 'POST':

				form5 = forms.form_remarks(request.POST)
				if form5.is_valid():
					obj3 = form5.save(commit=False)
					obj3.info = request.user
					obj3.year=y
					obj3.save()
					return HttpResponseRedirect(reverse('facform1:f_assistant4_final',args=(y,)))
			return render(request,'assistant_form4.html',{'form5':form5,'y':y})
		return render(request,'assistant_form4.html',{'form5':form5,'y':y})
	else:
		return HttpResponseRedirect('/invalid')
@login_required
def f_assistant_edit4(request,y):
	user=request.user
	obj=remarks.objects.filter(info=user).get(year=y)
	form2 = forms.form_remarks(instance=obj)
	if request.method == 'POST':
		# form1 = forms.form_User(request.POST)
		#form3 = form3(request.POST,request.FILES)
		form2 = forms.form_remarks(request.POST,request.FILES,instance=obj)

		if form2.is_valid():

			# sendme = User.objects.get(username=request.user)
			obj = form2.save(commit=False)
			obj.info = request.user
			obj.year=y
			obj.save()
			# sendme.doc_link  = 	form1.cleaned_data['doc_link']
			# sendme.save()

			return HttpResponseRedirect(reverse('facform1:f_assistant4_final',args=(y,)))
	return render(request,'assistant_form4.html',{'form5':form2,'y':y})

@login_required
def f_assistant4_final(request,y):
	user=request.user
	key2=rd.objects.filter(info=user).get(year=y)
	return render(request,'f_assistant4_final.html',{'key5':key2})

@login_required
def f_assistant3(request,y):
	if request.user.is_assistant_professor():
		form4 = forms.form_rd()
		if rd.objects.filter(info=request.user).filter(year=y).exists():
			return HttpResponseRedirect(reverse('facform1:assistant_form4',args=(y,)))
		else:
			if request.method == 'POST':
				form4 = forms.form_rd(request.POST,request.FILES)
				if form4.is_valid():
					obj2 = form4.save(commit=False)
					obj2.info = request.user
					obj2.year=y
					obj2.save()
					return HttpResponseRedirect(reverse('facform1:f_assistant3_final',args=(y,)))
			return render(request,'assistant_form3.html',{'form4':form4,'y':y})
		return render(request,'assistant_form3.html',{'form4':form4,'y':y})
	else:
		return HttpResponseRedirect('/invalid')

@login_required
def f_assistant_edit3(request,y):
	user=request.user
	obj=rd.objects.filter(info=user).get(year=y)
	form2 = forms.form_rd(instance=obj)
	if request.method == 'POST':
		# form1 = forms.form_User(request.POST)
		#form3 = form3(request.POST,request.FILES)
		form2 = forms.form_rd(request.POST,request.FILES,instance=obj)

		if form2.is_valid():

			# sendme = User.objects.get(username=request.user)
			obj = form2.save(commit=False)
			obj.info = request.user
			obj.year=y
			obj.save()
			# sendme.doc_link  = 	form1.cleaned_data['doc_link']
			# sendme.save()

			return HttpResponseRedirect(reverse('facform1:f_assistant3_final',args=(y,)))
	return render(request,'assistant_form3.html',{'form4':form2,'y':y})

@login_required
def f_assistant3_final(request,y):
	user=request.user
	key2=rd.objects.filter(info=user).get(year=y)
	return render(request,'f_assistant3_final.html',{'key4':key2,'y':y})

@login_required
def f_assistant2(request,y):
	print("Inside f_assistant2")
	user=request.user
	if request.user.is_assistant_professor():

		form3 = forms.form_feedbackTab()

		try:
			print("entered try")
			if feedbackTab.objects.filter(info=request.user).get(year=y):
				print("false")
				return HttpResponseRedirect(reverse('facform1:assistant_form3',args=(y,)))
		except:
			print("entered except")
			if request.method == 'POST':
				form3 = forms.form_feedbackTab(request.POST)
				print("hdsghj")
				if form3.is_valid():
					print("form is valid")
					obj1 = form3.save(commit=False)
					obj1.info = request.user
					obj1.year=y
					obj1.save()
					return HttpResponseRedirect(reverse('facform1:f_assistant2_final',args=(y,)))
			return render(request,'assistant_form2.html',{'form3':form3,'y':y})
		return render(request,'assistant_form2.html',{'form3':form3,'y':y})
	else:
		return HttpResponseRedirect('/invalid')

@login_required
def f_assistant_edit2(request,y):
	user=request.user
	data_final = User.objects.get(username=user)
	obj=feedbackTab.objects.filter(info=user).get(year=y)
	form2 = forms.form_feedbackTab(instance=obj)
	if request.method == 'POST':
		# form1 = forms.form_User(request.POST)
		#form3 = form3(request.POST,request.FILES)
		form2 = forms.form_feedbackTab(request.POST,request.FILES,instance=obj)

		if form2.is_valid():

			# sendme = User.objects.get(username=request.user)
			obj = form2.save(commit=False)
			obj.info = request.user
			obj.year=y
			obj.save()
			# sendme.doc_link  = 	form1.cleaned_data['doc_link']
			# sendme.save()

			return HttpResponseRedirect(reverse('facform1:f_assistant2_final',args=(y,)))
	return render(request,'assistant_form2.html',{'form3':form2,'y':y})

@login_required
def f_assistant2_final(request,y):
	user=request.user
	key2=feedbackTab.objects.filter(info=user).get(year=y)
	return render(request,'f_assistant2_final.html',{'key3':key2,'y':y})


@login_required
def f_assistant1(request,y):
	if request.user.is_assistant_professor():

		user = request.user
		print("inside f_assistant1")
		print(user)
		data_final = User.objects.get(username=user)
		# form1 = forms.form_User()
		form2 = forms.form_empDetailForm()
	#	form3=modelformset_factory(internships , fields = ('it_name','it_f','it_t'))
		try:
			if empDetailForm.objects.filter(info=user).get(year=y):
				return HttpResponseRedirect(reverse('facform1:assistant_form2',args=(y,)))

		except:
			if request.method == 'POST':
				# form1 = forms.form_User(request.POST)
				#form3 = form3(request.POST,request.FILES)
				form2 = forms.form_empDetailForm(request.POST,request.FILES)
				#f=empDetailForm.objects.filter(info=request.user).get(year="2018-19")

				if form2.is_valid():

					# sendme = User.objects.get(username=request.user)
					obj = form2.save(commit=False)
					obj.info = request.user
					obj.year=y
					obj.save()
					# sendme.doc_link  = 	form1.cleaned_data['doc_link']
					# sendme.save()
					return HttpResponseRedirect(reverse('facform1:f_assistant1_final',args=(y,)))
					#return HttpResponseRedirect("/assistant_form2/")
			return render(request,'assistant_form1.html',{'form2':form2,'info':data_final,'y':y})
		return render(request,'assistant_form1.html',{'form2':form2,'info':data_final,'y':y})#,'f':f})
	else:
		return HttpResponseRedirect('/invalid')

@login_required
def f_assistant_edit1(request,y):
	user=request.user
	data_final = User.objects.get(username=user)
	obj=empDetailForm.objects.filter(info=user).get(year=y)
	form2 = forms.form_empDetailForm(instance=obj)
	if request.method == 'POST':
		# form1 = forms.form_User(request.POST)
		#form3 = form3(request.POST,request.FILES)
		form2 = forms.form_empDetailForm(request.POST,request.FILES,instance=obj)

		if form2.is_valid():

			# sendme = User.objects.get(username=request.user)
			obj1 = form2.save(commit=False)
			obj1.info = request.user
			obj1.save()
			# sendme.doc_link  = 	form1.cleaned_data['doc_link']
			# sendme.save()

			return HttpResponseRedirect(reverse('facform1:f_assistant1_final',args=(y,)))
	return render(request,'assistant_form1.html',{'form2':form2,'info':data_final,'y':y})

@login_required
def f_assistant1_final(request,y):
	print("                                                                              Inside final view")
	print(y)
	user=request.user
	data = User.objects.get(username=user)

	key2=empDetailForm.objects.filter(info=user)
	print(key2)
	key2=key2.	get(year=str(y))
	return render(request,'f_assistant1_final.html',{'key1':data,'key2':key2,'y':y})



@login_required
def f_associate5(request,y):
	user=User.objects.get(username=request.user)
	print(user.designation)
	if user.is_associate_professor() or request.user.is_professor() or request.user.username == 'HOD_CIVIL':

		form6 = forms.form_conference()
		form7 = forms.form_journal()

		if conference.objects.filter(info=request.user).filter(year=y).exists():

			return HttpResponseRedirect(reverse('facform1:associate_preview',args=(y,)))
		else:
			if request.method == 'POST':
				sendme = User.objects.get(username=request.user)
				form6 = forms.form_conference(request.POST,request.FILES)
				form7 = forms.form_journal(request.POST,request.FILES)
				if form6.is_valid() and form7.is_valid():
					obj3 = form6.save(commit=False)
					obj4 = form7.save(commit=False)
					obj3.info = request.user
					obj4.info = request.user

					obj3.year=y
					obj4.year=y

					obj3.save()
					obj4.save()
					return HttpResponseRedirect(reverse('facform1:f_associate5_final',args=(y,)))
			else:
				print(form6.errors)
				print(form7.errors)
			return render(request,'associate_form5.html',{'form6':form6,'form7':form7,'y':y})
		return render(request,'associate_form5.html',{'form6':form6,'form7':form7,'y':y})
	else:
		return HttpResponseRedirect('/invalid')

@login_required
def f_associate_edit5(request,y):
	user=request.user
	obj1=conference.objects.filter(info=user).get(year=y)
	obj2=journal.objects.filter(info=user).get(year=y)
	form2= forms.form_conference(instance=obj1)
	form3=forms.form_journal(instance=obj2)
	if request.method == 'POST':
		# form1 = forms.form_User(request.POST)
		#form3 = form3(request.POST,request.FILES)
		form2 = forms.form_conference(request.POST,request.FILES,instance=obj1)
		form3=forms.form_journal(request.POST,request.FILES,instance=obj2)

		if form2.is_valid() and form3.is_valid():

			# sendme = User.objects.get(username=request.user)
			obj1 = form2.save(commit=False)
			obj1.info = request.user
			obj1.year=y
			obj1.save()
			obj2 = form3.save(commit=False)
			obj2.info = request.user
			obj2.year=y
			obj2.save()
			# sendme.doc_link  = 	form1.cleaned_data['doc_link']
			# sendme.save()
			return HttpResponseRedirect(reverse('facform1:f_associate5_final',args=(y,)))
	return render(request,'associate_form5.html',{'form6':form2,'form7':form3,'y':y})


	return HttpResponseRedirect(reverse('facform1:logout'))
@login_required
def f_associate5_final(request,y):
	user=request.user

	key6=conference.objects.filter(info=user).get(year=y)
	key7=journal.objects.filter(info=user).get(year=y)
	return render(request,'f_associate5_final.html',{'key6':key6,'key7':key7,'y':y})


@login_required
def f_associate4(request,y):
	if request.user.is_associate_professor() or request.user.is_professor() or request.user.username == 'HOD_CIVIL':
		form5 = forms.form_remarks()
		if remarks.objects.filter(info=request.user).filter(year=y).exists():

			return HttpResponseRedirect(reverse('facform1:associate_form5',args=(y,)))
		else:
			if request.method == 'POST':
				sendme = User.objects.get(username=request.user)
				form5 = forms.form_remarks(request.POST)
				if form5.is_valid():
					obj3 = form5.save(commit=False)
					obj3.info = request.user
					obj3.year=y
					obj3.save()
					return HttpResponseRedirect(reverse('facform1:f_associate4_final',args=(y,)))
			return render(request,'associate_form4.html',{'form5':form5,'y':y})
		return render(request,'associate_form4.html',{'form5':form5,'y':y})
	else:
		return HttpResponseRedirect('/invalid')

@login_required
def f_associate_edit4(request,y):
	user=request.user
	obj=remarks.objects.filter(info=user).get(year=y)
	form2 = forms.form_remarks(instance=obj)
	if request.method == 'POST':
		# form1 = forms.form_User(request.POST)
		#form3 = form3(request.POST,request.FILES)
		form2 = forms.form_remarks(request.POST,request.FILES,instance=obj)

		if form2.is_valid():

			# sendme = User.objects.get(username=request.user)
			obj = form2.save(commit=False)
			obj.info = request.user
			obj.year=y
			obj.save()
			# sendme.doc_link  = 	form1.cleaned_data['doc_link']
			# sendme.save()

			return HttpResponseRedirect(reverse('facform1:f_associate4_final',args=(y,)))
	return render(request,'associate_form4.html',{'form5':form2,'y':y})

@login_required
def f_associate4_final(request,y):
	user=request.user
	key2=remarks.objects.filter(info=user).get(year=y)
	return render(request,'f_associate4_final.html',{'key5':key2,'y':y})


@login_required
def f_associate3(request,y):
	if request.user.is_associate_professor() or request.user.is_professor() or request.user.username == 'HOD_CIVIL':
		form4 = forms.form_rd()
		if rd.objects.filter(info=request.user).filter(year=y).exists():

			return HttpResponseRedirect(reverse('facform1:associate_form4',args=(y,)))
		else:
			if request.method == 'POST':
				form4 = forms.form_rd(request.POST,request.FILES)
				if form4.is_valid():
					obj2 = form4.save(commit=False)
					obj2.info = request.user
					obj2.year=y
					obj2.save()
					return HttpResponseRedirect(reverse('facform1:f_associate3_final',args=(y,)))
			return render(request,'associate_form3.html',{'form4':form4,'y':y})
		return render(request,'associate_form3.html',{'form4':form4,'y':y})
	else:
		return HttpResponseRedirect('/invalid')
@login_required
def f_associate_edit3(request,y):
	user=request.user
	obj=rd.objects.filter(info=user).get(year=y)
	form2 = forms.form_rd(instance=obj)
	if request.method == 'POST':
		# form1 = forms.form_User(request.POST)
		#form3 = form3(request.POST,request.FILES)
		form2 = forms.form_rd(request.POST,request.FILES,instance=obj)

		if form2.is_valid():

			# sendme = User.objects.get(username=request.user)
			obj = form2.save(commit=False)
			obj.info = request.user
			obj.year=y
			obj.save()
			# sendme.doc_link  = 	form1.cleaned_data['doc_link']
			# sendme.save()

			return HttpResponseRedirect(reverse('facform1:f_associate3_final',args=(y,)))
	return render(request,'associate_form3.html',{'form4':form2,'y':y})

@login_required
def f_associate3_final(request,y):
	user=request.user
	key2=rd.objects.filter(info=user).get(year=y)
	return render(request,'f_associate3_final.html',{'key4':key2,'y':y})


@login_required
def f_associate2(request,y):
	if request.user.is_associate_professor() or request.user.is_professor() or request.user.username == 'HOD_CIVIL':
		form3 = forms.form_feedbackTab()

		if feedbackTab.objects.filter(info=request.user).filter(year=y).exists():

			return HttpResponseRedirect(reverse('facform1:associate_form3',args=(y,)))
		else:
			if request.method == 'POST':
				form3 = forms.form_feedbackTab(request.POST)
				if form3.is_valid():
					obj1 = form3.save(commit=False)
					obj1.info = request.user
					obj1.year=y
					obj1.save()
					return HttpResponseRedirect(reverse('facform1:f_associate2_final',args=(y,)))
			return render(request,'associate_form2.html',{'form3':form3,'y':y})

		return render(request,'associate_form2.html',{'form3':form3,'y':y})
	else:
		return HttpResponseRedirect('/invalid')

@login_required
def f_associate_edit2(request,y):
	user=request.user
	data_final = User.objects.get(username=user)
	obj=feedbackTab.objects.filter(info=user).get(year=y)
	form2 = forms.form_feedbackTab(instance=obj)
	if request.method == 'POST':
		# form1 = forms.form_User(request.POST)
		#form3 = form3(request.POST,request.FILES)
		form2 = forms.form_feedbackTab(request.POST,request.FILES,instance=obj)

		if form2.is_valid():

			# sendme = User.objects.get(username=request.user)
			obj = form2.save(commit=False)
			obj.info = request.user
			obj.year=y
			obj.save()
			# sendme.doc_link  = 	form1.cleaned_data['doc_link']
			# sendme.save()
			print("INSIDE f_assistant2  EDIT")
			return HttpResponseRedirect(reverse('facform1:f_associate2_final',args=(y,)))
	return render(request,'associate_form2.html',{'form3':form2,'y':y})

@login_required
def f_associate2_final(request,y):
	user=request.user
	key2=feedbackTab.objects.filter(info=user).get(year=y)
	return render(request,'f_associate2_final.html',{'key3':key2,'y':y})

@login_required
def f_associate1(request,y):
	if request.user.is_associate_professor() or request.user.is_professor() or request.user.username == 'HOD_CIVIL':
		user = request.user
		print(user)
		data_final = User.objects.get(username=user)
		# form1 = forms.form_User()
		form2 = forms.form_empDetailForm()
		if empDetailForm.objects.filter(info=user).filter(year=y).exists():
			return HttpResponseRedirect(reverse('facform1:associate_form2',args=(y,)))
		else:
			if request.method == 'POST':
				# form1 = forms.form_User(request.POST)
				form2 = forms.form_empDetailForm(request.POST,request.FILES)
				# print(form1.errors)
				print(form2.errors)
				if form2.is_valid():

					# sendme = User.objects.get(username=request.user)

					obj = form2.save(commit=False)
					obj.info = request.user
					obj.year=y
					obj.save()

					# sendme.doc_link  = 	form1.cleaned_data['doc_link']
					# sendme.save()

					return HttpResponseRedirect(reverse('facform1:f_associate1_final',args=(y,)))
			return render(request,'associate_form1.html',{'form2':form2,'info':data_final,'y':y})
		return render(request,'associate_form1.html',{'form2':form2,'info':data_final,'y':y})
	else:
		return HttpResponseRedirect('/invalid')


@login_required
def f_associate_edit1(request,y):
	user=request.user
	data_final = User.objects.get(username=user)
	obj=empDetailForm.objects.filter(info=user).get(year=y)
	form2 = forms.form_empDetailForm(instance=obj)
	if request.method == 'POST':
		# form1 = forms.form_User(request.POST)
		#form3 = form3(request.POST,request.FILES)
		form2 = forms.form_empDetailForm(request.POST,request.FILES,instance=obj)

		if form2.is_valid():

			# sendme = User.objects.get(username=request.user)
			obj = form2.save(commit=False)
			obj.info = request.user
			obj.save()
			# sendme.doc_link  = 	form1.cleaned_data['doc_link']
			# sendme.save()

			return HttpResponseRedirect(reverse('facform1:f_associate1_final',args=(y,)))
	return render(request,'associate_form1.html',{'form2':form2,'info':data_final,'y':y})

@login_required
def f_associate1_final(request,y):
	print("                                                                              Inside final view")
	print(y)
	user=request.user
	data = User.objects.get(username=user)

	key2=empDetailForm.objects.filter(info=user)
	print(key2)
	key2=key2.get(year=str(y))
	return render(request,'f_associate1_final.html',{'key1':data,'key2':key2,'y':y})



@login_required
def hod_form5(request,y):
	if request.user.is_hod():

		form6 = forms.form_conference()
		form7 = forms.form_journal()

		if conference.objects.filter(info=request.user).filter(year=y).exists():

			return HttpResponseRedirect(reverse('facform1:hod_preview',args=(y,)))
		else:
			if request.method == 'POST':
				sendme = User.objects.get(username=request.user)
				form6 = forms.form_conference(request.POST,request.FILES)
				form7 = forms.form_journal(request.POST,request.FILES)
				if form6.is_valid() and form7.is_valid():
					obj3 = form6.save(commit=False)
					obj4 = form7.save(commit=False)
					obj3.info = request.user
					obj4.info = request.user

					obj4.year=y
					obj3.year=y

					obj3.save()
					obj4.save()

					return HttpResponseRedirect(reverse('facform1:hod_form5_final',args=(y,)))
			return render(request,'hod_form5.html',{'form6':form6,'form7':form7,'y':y})
		return render(request,'hod_form5.html',{'form6':form6,'form7':form7,'y':y})
	else:
		return HttpResponseRedirect('/invalid')


@login_required
def hod_form5_edit(request,y):
	user=request.user
	obj1=conference.objects.filter(info=user).get(year=y)
	obj2=journal.objects.filter(info=user).get(year=y)
	form2= forms.form_conference(instance=obj1)
	form3=forms.form_journal(instance=obj2)
	if request.method == 'POST':
		# form1 = forms.form_User(request.POST)
		#form3 = form3(request.POST,request.FILES)
		form2 = forms.form_conference(request.POST,request.FILES,instance=obj1)
		form3=forms.form_journal(request.POST,request.FILES,instance=obj2)

		if form2.is_valid() and form3.is_valid():

			# sendme = User.objects.get(username=request.user)
			obj1 = form2.save(commit=False)
			obj1.info = request.user
			obj1.year=y
			obj1.save()
			obj2 = form3.save(commit=False)
			obj2.info = request.user
			obj2.year=y
			obj2.save()
			# sendme.doc_link  = 	form1.cleaned_data['doc_link']
			# sendme.save()

			return HttpResponseRedirect(reverse('facform1:hod_form5_final',args=(y,)))
	return render(request,'hod_form5.html',{'form6':form2,'form7':form3,'y':y})

	return HttpResponseRedirect(reverse('facform1:logout'))
@login_required
def hod_form5_final(request,y):
	user=request.user

	key6=conference.objects.filter(info=user).get(year=y)
	key7=journal.objects.filter(info=user).get(year=y)
	return render(request,'hod_form5_final.html',{'key6':key6,'key7':key7,'y':y})




@login_required
def hod_form4(request,y):
	if request.user.is_hod():
		form5 = forms.form_remarks()
		if remarks.objects.filter(info=request.user).filter(year=y).exists():

			return HttpResponseRedirect(reverse('facform1:hod_form5',args=(y,)))
		else:
			if request.method == 'POST':
				form5 = forms.form_remarks(request.POST)
				if form5.is_valid():
					obj3 = form5.save(commit=False)
					obj3.info = request.user
					obj3.year=y
					obj3.save()
					return HttpResponseRedirect(reverse('facform1:hod_form4_final',args=(y,)))
			return render(request,'hod_form4.html',{'form5':form5,'y':y})
		return render(request,'hod_form4.html',{'form5':form5,'y':y})
	else:
		return HttpResponseRedirect('/invalid')



@login_required
def hod_form4_edit(request,y):
	user=request.user
	obj=remarks.objects.filter(info=user).get(year=y)
	form2 = forms.form_remarks(instance=obj)
	if request.method == 'POST':
		# form1 = forms.form_User(request.POST)
		#form3 = form3(request.POST,request.FILES)
		form2 = forms.form_remarks(request.POST,request.FILES,instance=obj)

		if form2.is_valid():

			# sendme = User.objects.get(username=request.user)
			obj = form2.save(commit=False)
			obj.info = request.user
			obj.year=y
			obj.save()
			# sendme.doc_link  = 	form1.cleaned_data['doc_link']
			# sendme.save()

			return HttpResponseRedirect(reverse('facform1:hod_form4_final',args=(y,)))
	return render(request,'hod_form4.html',{'form5':form2,'y':y})

@login_required
def hod_form4_final(request,y):
	user=request.user
	key2=remarks.objects.filter(info=user).get(year=y)
	return render(request,'hod_form4_final.html',{'key5':key2,'y':y})





@login_required
def hod_form3(request,y):
	if request.user.is_hod():
		form4 = forms.form_rd()
		if rd.objects.filter(info=request.user).filter(year=y).exists():

			return HttpResponseRedirect(reverse('facform1:hod_form4',args=(y,)))
		else:
			if request.method == 'POST':
				form4 = forms.form_rd(request.POST,request.FILES)
				if form4.is_valid():
					obj2 = form4.save(commit=False)
					obj2.info = request.user
					obj2.year=y
					obj2.save()
					return HttpResponseRedirect(reverse('facform1:hod_form3_final',args=(y,)))
			return render(request,'hod_form3.html',{'form4':form4,'y':y})
		return render(request,'hod_form3.html',{'form4':form4,'y':y})
	else:
		return HttpResponseRedirect('/invalid')

@login_required
def hod_form3_edit(request,y):
	user=request.user
	obj=rd.objects.filter(info=user).get(year=y)
	form2 = forms.form_rd(instance=obj)
	if request.method == 'POST':
		# form1 = forms.form_User(request.POST)
		#form3 = form3(request.POST,request.FILES)
		form2 = forms.form_rd(request.POST,request.FILES,instance=obj)

		if form2.is_valid():

			# sendme = User.objects.get(username=request.user)
			obj = form2.save(commit=False)
			obj.info = request.user
			obj.year=y
			obj.save()
			# sendme.doc_link  = 	form1.cleaned_data['doc_link']
			# sendme.save()

			return HttpResponseRedirect(reverse('facform1:hod_form3_final',args=(y,)))
	return render(request,'hod_form3.html',{'form4':form2,'y':y})

@login_required
def hod_form3_final(request,y):
	user=request.user
	key2=rd.objects.filter(info=user).get(year=y)
	return render(request,'hod_form3_final.html',{'key4':key2,'y':y})


@login_required
def hod_form2(request,y):
	if request.user.is_hod():
		form3 = forms.form_feedbackTab()

		if feedbackTab.objects.filter(info=request.user).filter(year=y).exists():

			return HttpResponseRedirect(reverse('facform1:hod_form3',args=(y,)))
		else:
			if request.method == 'POST':
				form3 = forms.form_feedbackTab(request.POST)
				if form3.is_valid():
					obj1 = form3.save(commit=False)
					obj1.info = request.user
					obj1.year=y
					obj1.save()
					return HttpResponseRedirect(reverse('facform1:hod_form2_final',args=(y,)))
			return render(request,'hod_form2.html',{'form3':form3,'y':y})

		return render(request,'hod_form2.html',{'form3':form3,'y':y})
	else:
		return HttpResponseRedirect('/invalid')


@login_required
def hod_form2_edit(request,y):
	user=request.user
	data_final = User.objects.get(username=user)
	obj=feedbackTab.objects.filter(info=user).get(year=y)
	form2 = forms.form_feedbackTab(instance=obj)
	if request.method == 'POST':
		# form1 = forms.form_User(request.POST)
		#form3 = form3(request.POST,request.FILES)
		form2 = forms.form_feedbackTab(request.POST,request.FILES,instance=obj)

		if form2.is_valid():

			# sendme = User.objects.get(username=request.user)
			obj = form2.save(commit=False)
			obj.info = request.user
			obj.year=y
			obj.save()
			# sendme.doc_link  = 	form1.cleaned_data['doc_link']
			# sendme.save()

			return HttpResponseRedirect(reverse('facform1:hod_form2_final',args=(y,)))
	return render(request,'hod_form2.html',{'form3':form2,'y':y})

@login_required
def hod_form2_final(request,y):
	user=request.user
	key2=feedbackTab.objects.filter(info=user).get(year=y)
	return render(request,'hod_form2_final.html',{'key3':key2,'y':y})



@login_required
def hod_form1(request,y):
	if request.user.is_hod():

		user = request.user
		print(user)
		data_final = User.objects.get(username=user)
		# form1 = forms.form_User()
		form2 = forms.form_empDetailForm()

		if empDetailForm.objects.filter(info=user).exists():
			print("Hey")
			return HttpResponseRedirect(reverse('facform1:hod_form2',args=(y,)))

		else:
			if request.method == 'POST':
				print("HEY2")
				# form1 = forms.form_User(request.POST)
				form2 = forms.form_empDetailForm(request.POST,request.FILES)


				if form2.is_valid():
					print("HOD FORM 1 IS VALID ")
					# sendme = User.objects.get(username=request.user)

					obj = form2.save(commit=False)
					obj.info = request.user
					obj.year=y
					obj.save()

					print("Redirecting to form 1 review page")
					return HttpResponseRedirect(reverse('facform1:hod_form1_final',args=(y,)))
			return render(request,'hod_form1.html',{'form2':form2,'info':data_final,'y':y})
		return render(request,'hod_form1.html',{'form2':form2,'info':data_final,'y':y})

	else:
		return HttpResponseRedirect('/invalid')


@login_required
def hod_form1_edit(request,y):
	user=request.user
	data_final = User.objects.get(username=user)
	obj=empDetailForm.objects.filter(info=user).get(year=y)
	form2 = forms.form_empDetailForm(instance=obj)
	if request.method == 'POST':
		# form1 = forms.form_User(request.POST)
		#form3 = form3(request.POST,request.FILES)
		form2 = forms.form_empDetailForm(request.POST,request.FILES,instance=obj)

		if form2.is_valid():

			# sendme = User.objects.get(username=request.user)
			obj = form2.save(commit=False)
			obj.info = request.user
			obj.year=y
			obj.save()
			# sendme.doc_link  = 	form1.cleaned_data['doc_link']
			# sendme.save()

			return HttpResponseRedirect(reverse('facform1:hod_form1_final',args=(y,)))
	return render(request,'hod_form1.html',{'form2':form2,'info':data_final,'y':y})


@login_required
def hod_form1_final(request,y):
	print("                                                                              Inside final view of first form ")
	print(y)
	user=request.user
	data = User.objects.get(username=user)
	key2=empDetailForm.objects.filter(info=user)
	print(key2)
	key2=key2.get(year=str(y))
	return render(request,'hod_form1_final.html',{'key1':data,'key2':key2,'y':y})







@login_required
def assistant_preview(request,y):
	if request.user.is_assistant_professor():

		name = request.user
		n=new.objects.filter(info=name).get(year__year=y)
		data1 = User.objects.get(username=name)
		data2 = empDetailForm.objects.filter(info=name).get(year=y)
		data3 = feedbackTab.objects.filter(info=name).get(year=y)
		data4 = rd.objects.filter(info=name).get(year=y)
		data5 = remarks.objects.filter(info=name).get(year=y)
		data6 = conference.objects.filter(info=name).get(year=y)
		data7 = journal.objects.filter(info=name).get(year=y)

		if new.objects.filter(info__username=name).filter(year__year=y).filter(hod_status=True):
			data8 = remarks1.objects.get(info__username=name)
		else:
			data8 = []
		if new.objects.filter(info__username=name).filter(year__year=y).filter(principal_status=True):
			data9 = remarks2.objects.get(info__username=name)
		else:
			data9 = []


		context1 = {
		"key1":data1,
		"key2":data2,
		"key3":data3,
		"key4":data4,
		"key5":data5,
		"key6":data6,
		"key7":data7,
		"key8":data8,
		"key9":data9,
		"n":n
		}

		return render(request,'assistant_preview.html',context=context1)
	else:
		return HttpResponseRedirect('/invalid')







@login_required
def associate_preview(request,y):
	name = request.user
	n=new.objects.filter(info=name).get(year__year=y)
	data1 = User.objects.get(username=name)
	data2 = empDetailForm.objects.filter(info=name).get(year=y);
	data3 = feedbackTab.objects.filter(info=name).get(year=y)
	data4 = rd.objects.filter(info=name).get(year=y)
	data5 = remarks.objects.filter(info=name).get(year=y)
	data6 = conference.objects.filter(info=name).get(year=y)
	data7 = journal.objects.filter(info=name).get(year=y)

	if new.objects.filter(info__username=name).filter(hod_status=True):
		data8 = remarks1.objects.get(info__username=name)
	else:
		data8 = []
	if new.objects.filter(info__username=name).filter(principal_status=True):
		data9 = remarks2.objects.get(info__username=name)
	else:
		data9 = []

	context1 = {
	"n":n,
	"key1":data1,
	"key2":data2,
	"key3":data3,
	"key4":data4,
	"key5":data5,
	"key6":data6,
	"key7":data7,
	"key8":data8,
	"key9":data9,
	}

	return render(request,'associate_preview.html',context=context1)

@login_required
def hod_preview(request,y):
	name = request.user
	n=new.objects.filter(info=name).get(year__year=y)
	data1 = User.objects.get(username=name)
	data2 = empDetailForm.objects.filter(info=name).get(year=y);
	data3 = feedbackTab.objects.filter(info=name).get(year=y);
	data4 = rd.objects.filter(info=name).get(year=y);
	data5 = remarks.objects.filter(info=name).get(year=y);
	data6 = conference.objects.filter(info=name).get(year=y);
	data7 = journal.objects.filter(info=name).get(year=y);
	if new.objects.filter(info__username=name).filter(principal_status=True):
		data8 = remarks2.objects.get(info__username=name)
	else:
		data8 = []

	context1 = {
	"key1":data1,
	"key2":data2,
	"key3":data3,
	"key4":data4,
	"key5":data5,
	"key6":data6,
	"key7":data7,
	"key8":data8,
	"n":n
	}

	return render(request,'hod_preview.html',context=context1)
